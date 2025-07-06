from typing import Any, Dict, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

from .base import DocumentParser
from ..layout_analyser import LayoutAnalyzer
from ..types import BoundingBox, DocumentElement, DocumentStructure, ElementType, ParsedDocument


class PositionalExcelParser(DocumentParser):
    """Enhanced Excel parser that preserves worksheet structure and cell positioning"""
    
    def can_parse(self, file_path: str) -> bool:
        return file_path.lower().endswith(('.xlsx', '.xls', '.xlsm'))
    
    def parse(self, file_path: str) -> ParsedDocument:
        elements = []
        metadata = {}
        errors = []
        page_dimensions = {}
        
        try:
            # Use openpyxl for better cell positioning and formatting info
            if file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xlsm'):
                workbook = load_workbook(file_path, data_only=True)
                
                metadata = {
                    'worksheets': len(workbook.worksheets),
                    'sheet_names': workbook.sheetnames,
                    'active_sheet': workbook.active.title if workbook.active else None
                }
                
                # Process each worksheet as a separate "page"
                for page_num, worksheet in enumerate(workbook.worksheets):
                    sheet_metadata = {
                        'sheet_name': worksheet.title,
                        'max_row': worksheet.max_row,
                        'max_column': worksheet.max_column,
                        'dimensions': worksheet.dimensions
                    }
                    
                    # Calculate page dimensions based on used range
                    max_col_width = worksheet.max_column * 100  # Approximate column width
                    max_row_height = worksheet.max_row * 20     # Approximate row height
                    page_dimensions[page_num] = (max_col_width, max_row_height)
                    
                    # Extract merged cell ranges
                    merged_ranges = [str(merged_range) for merged_range in worksheet.merged_cells.ranges]
                    
                    # Process cells with positioning
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                # Calculate cell position
                                col_offset = (cell.column - 1) * 100  # Approximate column width
                                row_offset = (cell.row - 1) * 20      # Approximate row height
                                
                                bbox = BoundingBox(
                                    x0=col_offset,
                                    y0=row_offset,
                                    x1=col_offset + 100,
                                    y1=row_offset + 20,
                                    page=page_num
                                )
                                
                                # Determine element type based on cell properties
                                element_type = self._classify_cell_element(cell, worksheet)
                                
                                # Get cell styling information
                                style_info = self._extract_cell_style(cell)
                                
                                element = DocumentElement(
                                    content=str(cell.value),
                                    element_type=element_type,
                                    bbox=bbox,
                                    style=style_info,
                                    metadata={
                                        'cell_address': cell.coordinate,
                                        'column': cell.column,
                                        'row': cell.row,
                                        'column_letter': get_column_letter(cell.column),
                                        'sheet_name': worksheet.title,
                                        'data_type': str(type(cell.value).__name__),
                                        'is_merged': cell.coordinate in [str(range) for range in worksheet.merged_cells.ranges]
                                    }
                                )
                                elements.append(element)
                    
                    # Detect and extract table structures
                    tables = self._detect_excel_tables(worksheet, page_num)
                    elements.extend(tables)
                    
                    # Extract headers (first row with bold formatting or first row of data)
                    headers = self._extract_headers(worksheet, page_num)
                    elements.extend(headers)
                
                workbook.close()
                
            else:
                # Fallback to pandas for .xls files
                excel_data = pd.read_excel(file_path, sheet_name=None)
                
                metadata = {
                    'worksheets': len(excel_data),
                    'sheet_names': list(excel_data.keys())
                }
                
                for page_num, (sheet_name, df) in enumerate(excel_data.items()):
                    page_dimensions[page_num] = (len(df.columns) * 100, len(df) * 20)
                    
                    # Process DataFrame as positioned elements
                    for row_idx, row in df.iterrows():
                        for col_idx, value in enumerate(row):
                            if pd.notna(value):
                                bbox = BoundingBox(
                                    x0=col_idx * 100,
                                    y0=row_idx * 20,
                                    x1=(col_idx + 1) * 100,
                                    y1=(row_idx + 1) * 20,
                                    page=page_num
                                )
                                
                                element = DocumentElement(
                                    content=str(value),
                                    element_type=ElementType.TEXT,
                                    bbox=bbox,
                                    metadata={
                                        'cell_address': f"{get_column_letter(col_idx + 1)}{row_idx + 1}",
                                        'column': col_idx + 1,
                                        'row': row_idx + 1,
                                        'column_name': df.columns[col_idx],
                                        'sheet_name': sheet_name
                                    }
                                )
                                elements.append(element)
                
        except Exception as e:
            errors.append(f"Excel parsing error: {str(e)}")
        
        reading_order = LayoutAnalyzer.determine_reading_order(elements)
        
        structure = DocumentStructure(
            elements=elements,
            reading_order=reading_order,
            page_dimensions=page_dimensions
        )
        
        return ParsedDocument(
            structure=structure,
            metadata=metadata,
            file_path=file_path,
            file_type="excel",
            parser_used="PositionalExcelParser",
            errors=errors if errors else None
        )
    
    def _classify_cell_element(self, cell, worksheet) -> ElementType:
        """Classify cell element based on position and formatting"""
        # Check if it's a header (first row or bold)
        if cell.row == 1:
            return ElementType.HEADING
        
        # Check if it's part of a formula
        if hasattr(cell, 'formula') and cell.formula:
            return ElementType.FORMULA
        
        # Check if it's a number vs text
        if isinstance(cell.value, (int, float)):
            return ElementType.TEXT
        elif isinstance(cell.value, str):
            # Check if it looks like a heading (short, capitalized)
            if len(cell.value) < 50 and cell.value.isupper():
                return ElementType.HEADING
            return ElementType.TEXT
        
        return ElementType.TEXT
    
    def _extract_cell_style(self, cell) -> Dict[str, Any]:
        """Extract styling information from cell"""
        style = {}
        
        if hasattr(cell, 'font') and cell.font:
            style['font_name'] = cell.font.name
            style['font_size'] = cell.font.size
            style['bold'] = cell.font.bold
            style['italic'] = cell.font.italic
            style['color'] = str(cell.font.color) if cell.font.color else None
        
        if hasattr(cell, 'fill') and cell.fill:
            style['fill_color'] = str(cell.fill.start_color) if cell.fill.start_color else None
        
        if hasattr(cell, 'alignment') and cell.alignment:
            style['horizontal_alignment'] = cell.alignment.horizontal
            style['vertical_alignment'] = cell.alignment.vertical
        
        return style
    
    def _detect_excel_tables(self, worksheet, page_num: int) -> List[DocumentElement]:
        """Detect and extract table structures from worksheet"""
        tables = []
        
        # Simple table detection: look for rectangular regions of data
        data_cells = [(cell.row, cell.column) for row in worksheet.iter_rows() 
                     for cell in row if cell.value is not None]
        
        if not data_cells:
            return tables
        
        # Find contiguous rectangular regions
        min_row = min(cell[0] for cell in data_cells)
        max_row = max(cell[0] for cell in data_cells)
        min_col = min(cell[1] for cell in data_cells)
        max_col = max(cell[1] for cell in data_cells)
        
        # Create table element for the main data region
        bbox = BoundingBox(
            x0=(min_col - 1) * 100,
            y0=(min_row - 1) * 20,
            x1=max_col * 100,
            y1=max_row * 20,
            page=page_num
        )
        
        # Extract table data
        table_data = []
        for row_num in range(min_row, max_row + 1):
            row_data = []
            for col_num in range(min_col, max_col + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                row_data.append(str(cell.value) if cell.value is not None else "")
            table_data.append(row_data)
        
        table_content = "\n".join(["\t".join(row) for row in table_data])
        
        table_element = DocumentElement(
            content=table_content,
            element_type=ElementType.TABLE,
            bbox=bbox,
            metadata={
                'table_data': table_data,
                'sheet_name': worksheet.title,
                'dimensions': f"{min_row}:{max_row},{min_col}:{max_col}"
            }
        )
        tables.append(table_element)
        
        return tables
    
    def _extract_headers(self, worksheet, page_num: int) -> List[DocumentElement]:
        """Extract header elements from worksheet"""
        headers = []
        
        # Check first row for headers
        first_row = list(worksheet.iter_rows(min_row=1, max_row=1))[0]
        
        for cell in first_row:
            if cell.value is not None:
                bbox = BoundingBox(
                    x0=(cell.column - 1) * 100,
                    y0=0,
                    x1=cell.column * 100,
                    y1=20,
                    page=page_num
                )
                
                header_element = DocumentElement(
                    content=str(cell.value),
                    element_type=ElementType.HEADING,
                    bbox=bbox,
                    style=self._extract_cell_style(cell),
                    metadata={
                        'cell_address': cell.coordinate,
                        'is_header': True,
                        'sheet_name': worksheet.title
                    }
                )
                headers.append(header_element)
        
        return headers